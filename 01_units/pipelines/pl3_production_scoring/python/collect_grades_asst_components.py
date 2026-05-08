from __future__ import annotations

"""Aggregate component-level grade exports for an umbrella assignment.

This script scans assignment subdirectories under ``--assignment-root`` and
collects component gradebook outputs produced by stage02. It builds merged
umbrella outputs for both the regular grade files and, when present, the
``_SSID`` variants.

Inputs
------
- ``--assignment-root``
    Umbrella directory that contains component subdirectories with
    ``pipeline_paths.json``.
- Per-component stage02 outputs discovered through each component's
    ``pipeline_paths.json``:
    - ``.../04_gradebook_populated/*_Grades_iterXX.csv``
    - optional sibling ``.../*_Grades_iterXX_SSID.csv``
- Optional weights file at
    ``<assignment-root>/04_grades_release/<token>_weights.json``.

Behavior
--------
- Resolves each component's populated gradebook CSV path from config.
- Validates identity columns and enforces identity-set consistency across
    all discovered components in a merge set.
- Writes merged umbrella outputs with two component blocks:
    - raw component grade/max/feedback columns
    - weighted component grade/max columns (plus per-component weight)
- Writes merged umbrella outputs for the regular set.
- If SSID component files are discovered, writes a second merged SSID set.
- Reports skipped components when expected CSV files are missing.

Outputs
-------
Under ``<assignment-root>/04_grades_release/``:
- ``<token>_component_grade_feedback_merged.xlsx``
- ``<token>_Grades_iterXX.csv``
- (when SSID sources exist) ``<token>_component_grade_feedback_merged_SSID.xlsx``
- (when SSID sources exist) ``<token>_Grades_iterXX_SSID.csv``
"""

import argparse
import csv
import json
import math
import re
import statistics
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.utils import get_column_letter


IDENTITY_COLUMNS = ["Identifier", "Full name", "Email address"]
GRADE_COLUMN = "Grade"
MAX_GRADE_COLUMN = "Maximum Grade"
FEEDBACK_COLUMN_CANDIDATES = ["Feedback comments", "Feedback"]
SEPARATOR_COLUMN = "."
WEIGHTED_SCORE_COLUMN = "submission_numeric_score"
MAX_SCORE_COLUMN = "submission_max_numeric_score"
AGGREGATED_FEEDBACK_COLUMN = "Feedback comments"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge component-level gradebook outputs for one assignment umbrella directory."
    )
    parser.add_argument(
        "--assignment-root",
        type=Path,
        required=True,
        help="Assignment umbrella directory, such as ./PPP or ./PPS1-umbrella.",
    )
    return parser.parse_args()


def load_pipeline_paths_config(assignment_dir: Path) -> dict[str, object] | None:
    candidates = [
        assignment_dir / "pipeline_paths.json",
        assignment_dir / "pipeline_paths.jsom",
    ]
    for path in candidates:
        if path.exists() and path.is_file():
            return json.loads(path.read_text(encoding="utf-8"))
    return None


def resolve_output_csv_from_config(assignment_dir: Path, config: dict[str, object]) -> Path:
    layer4_prod = config.get("layer4_prod")
    if not isinstance(layer4_prod, dict):
        raise ValueError(f"Missing layer4_prod block in {assignment_dir / 'pipeline_paths.json'}")

    release = layer4_prod.get("release", {})
    if not isinstance(release, dict):
        release = {}
    iteration = str(release.get("iteration", "")).strip()
    filename_version = str(release.get("filename_version", "")).strip()

    populate = layer4_prod.get("populate_gradebook_export")
    if not isinstance(populate, dict):
        raise ValueError(
            f"Missing layer4_prod.populate_gradebook_export block in {assignment_dir / 'pipeline_paths.json'}"
        )

    output = populate.get("output")
    if not isinstance(output, dict):
        raise ValueError(
            f"Missing layer4_prod.populate_gradebook_export.output block in {assignment_dir / 'pipeline_paths.json'}"
        )

    def expand(value: str) -> str:
        expanded = value
        if iteration:
            expanded = expanded.replace("{iteration}", iteration)
        if filename_version:
            expanded = expanded.replace("{filename_version}", filename_version)
        return expanded

    base = expand(str(output.get("base", "") or ""))
    dir_part = expand(str(output.get("dir", "") or ""))
    run = expand(str(output.get("run", "") or ""))
    subdir = expand(str(output.get("run_relative_subdir", "") or ""))
    filename = expand(str(output.get("filename", "") or ""))
    if not filename:
        raise ValueError(
            f"Missing output.filename in layer4_prod.populate_gradebook_export.output for {assignment_dir.name}"
        )

    # pipeline_paths values are rooted under assignment_dir/01_pipelines/
    root = assignment_dir / "01_pipelines"
    csv_path = (root / base / dir_part / run / subdir / filename).resolve()
    return csv_path


def read_grade_rows(csv_path: Path) -> tuple[list[str], list[dict[str, str]], str, str]:
    if not csv_path.exists() or not csv_path.is_file():
        raise FileNotFoundError(f"Expected gradebook CSV not found: {csv_path}")

    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError(f"CSV has no header: {csv_path}")
        rows = [dict(row) for row in reader]
        headers = list(reader.fieldnames)

    for column in IDENTITY_COLUMNS + [GRADE_COLUMN]:
        if column not in headers:
            raise ValueError(f"Missing required column '{column}' in {csv_path}")

    if MAX_GRADE_COLUMN not in headers:
        raise ValueError(f"Missing required column '{MAX_GRADE_COLUMN}' in {csv_path}")

    feedback_column = next((name for name in FEEDBACK_COLUMN_CANDIDATES if name in headers), "")
    if not feedback_column:
        raise ValueError(
            f"Missing feedback column in {csv_path}. Expected one of: {FEEDBACK_COLUMN_CANDIDATES}"
        )

    return headers, rows, feedback_column, MAX_GRADE_COLUMN


def identity_key(row: dict[str, str]) -> tuple[str, str, str]:
    return tuple((row.get(column, "") or "").strip() for column in IDENTITY_COLUMNS)


def validate_and_index_identity_rows(
    source_name: str,
    rows: list[dict[str, str]],
) -> dict[tuple[str, str, str], dict[str, str]]:
    index: dict[tuple[str, str, str], dict[str, str]] = {}
    for row in rows:
        key = identity_key(row)
        if key in index:
            raise ValueError(f"Duplicate identity row detected in {source_name}: {key}")
        index[key] = row
    return index


def aggregate_feedback_for_identity(
    key: tuple[str, str, str],
    source_data: list[tuple[str, dict[tuple[str, str, str], dict[str, str]], str, str]],
    component_weights: dict[str, str],
) -> str:
    values: list[str] = []
    for source_name, rows_by_id, feedback_column, _max_grade_column in source_data:
        raw_text = (rows_by_id[key].get(feedback_column, "") or "").strip()
        text = scale_feedback_score_pairs(raw_text, component_weights.get(source_name, ""))
        if text:
            values.append(text)
    return "\n\n".join(values)


def load_component_weights(weights_path: Path) -> dict[str, str]:
    """Load component weights from JSON.

    Supported shapes:
    - {"PPS1E1": 0.2, "PPS1E21": 0.1}
    - {"weights": {"PPS1E1": 0.2, ...}}
    Keys may include or omit the "Weight_" prefix.
    """
    if not weights_path.is_file():
        raise FileNotFoundError(f"Required weights file not found: {weights_path}")

    raw_text = weights_path.read_text(encoding="utf-8").strip()
    if not raw_text:
        raise ValueError(f"Weights file is empty: {weights_path}")

    try:
        payload = json.loads(raw_text)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Weights file is not valid JSON: {weights_path}") from exc

    if not isinstance(payload, dict):
        raise ValueError(f"Weights file must contain a JSON object: {weights_path}")

    candidate = payload.get("weights")
    if isinstance(candidate, dict):
        source = candidate
    else:
        source = payload

    result: dict[str, str] = {}
    for key, value in source.items():
        if not isinstance(key, str):
            continue
        normalized_key = key.removeprefix("Weight_")
        if isinstance(value, dict):
            nested_weight = value.get("weight")
            if nested_weight is None:
                continue
            result[normalized_key] = str(nested_weight)
        else:
            result[normalized_key] = str(value)
    if not result:
        raise ValueError(
            f"No component weights found in required weights file: {weights_path}"
        )

    return result


def _parse_float(value: str) -> float | None:
    text = (value or "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def weighted_sum_and_denominator(
    grade_values: list[str],
    weight_values: list[str],
) -> tuple[str, str]:
    total = 0.0
    denominator = 0.0
    terms = 0
    for grade_raw, weight_raw in zip(grade_values, weight_values):
        grade = _parse_float(grade_raw)
        weight = _parse_float(weight_raw)
        if grade is None or weight is None:
            continue
        total += grade * weight
        denominator += weight
        terms += 1

    if terms == 0:
        return "", ""
    return f"{total:.2f}", f"{denominator:.2f}"


def scale_value_by_weight(value_raw: str, weight_raw: str) -> float | None:
    value = _parse_float(value_raw)
    if value is None:
        return None
    weight = _parse_float(weight_raw)
    if weight is None:
        return value
    return value * weight


def _format_number(value: float) -> str:
    text = f"{value:.2f}".rstrip("0").rstrip(".")
    return text if text else "0"


def scale_feedback_score_pairs(feedback_text: str, weight_raw: str) -> str:
    """Scale '(x / y)', 'x/y', and '(x)' numeric scores inside feedback text by component weight."""
    text = (feedback_text or "").strip()
    if not text:
        return ""

    weight = _parse_float(weight_raw)
    if weight is None:
        return text

    paren_pair_pattern = re.compile(r"\(\s*(-?\d+(?:\.\d+)?)\s*/\s*(-?\d+(?:\.\d+)?)\s*\)")
    bare_pair_pattern = re.compile(
        r"(?<!\()\b(-?\d+(?:\.\d+)?)(\s*/\s*)(-?\d+(?:\.\d+)?)\b(?!\))"
    )
    paren_value_pattern = re.compile(r"\(\s*(-?\d+(?:\.\d+)?)\s*\)")

    def _replace_paren_pair(match: re.Match[str]) -> str:
        numerator = float(match.group(1)) * weight
        denominator = float(match.group(2)) * weight
        return f"({_format_number(numerator)} / {_format_number(denominator)})"

    def _replace_bare_pair(match: re.Match[str]) -> str:
        numerator = float(match.group(1)) * weight
        separator = match.group(2)
        denominator = float(match.group(3)) * weight
        return f"{_format_number(numerator)}{separator}{_format_number(denominator)}"

    def _replace_paren_value(match: re.Match[str]) -> str:
        value = float(match.group(1)) * weight
        return f"({_format_number(value)})"

    scaled = paren_pair_pattern.sub(_replace_paren_pair, text)
    scaled = bare_pair_pattern.sub(_replace_bare_pair, scaled)
    scaled = paren_value_pattern.sub(_replace_paren_value, scaled)
    return scaled


def _compute_weighted_total(values: list[str], weights: list[str]) -> float | None:
    total = 0.0
    terms = 0
    for value_raw, weight_raw in zip(values, weights):
        weighted_value = scale_value_by_weight(value_raw, weight_raw)
        if weighted_value is None:
            continue
        total += weighted_value
        terms += 1
    if terms == 0:
        return None
    return total


def _quartiles(values: list[float]) -> tuple[float, float]:
    if not values:
        return 0.0, 0.0
    if len(values) == 1:
        return values[0], values[0]
    q1, _, q3 = statistics.quantiles(values, n=4, method="inclusive")
    return q1, q3


def _append_metric_rows(sheet, start_row: int, title: str, values: list[float]) -> int:
    sheet.cell(row=start_row, column=1).value = title
    start_row += 1
    headers = ["metric", "value"]
    for idx, header in enumerate(headers, start=1):
        sheet.cell(row=start_row, column=idx).value = header
    start_row += 1

    if not values:
        rows = [
            ("count", 0),
            ("mean", ""),
            ("median", ""),
            ("stdev", ""),
            ("min", ""),
            ("q1", ""),
            ("q3", ""),
            ("max", ""),
        ]
    else:
        sorted_values = sorted(values)
        q1, q3 = _quartiles(sorted_values)
        rows = [
            ("count", len(sorted_values)),
            ("mean", round(statistics.mean(sorted_values), 4)),
            ("median", round(statistics.median(sorted_values), 4)),
            ("stdev", round(statistics.stdev(sorted_values), 4) if len(sorted_values) > 1 else 0.0),
            ("min", round(sorted_values[0], 4)),
            ("q1", round(q1, 4)),
            ("q3", round(q3, 4)),
            ("max", round(sorted_values[-1], 4)),
        ]

    for metric, value in rows:
        sheet.cell(row=start_row, column=1).value = metric
        sheet.cell(row=start_row, column=2).value = value
        start_row += 1

    return start_row + 1


def _metric_row_values(values: list[float]) -> dict[str, float | int | str]:
    if not values:
        return {
            "count": 0,
            "mean": "",
            "median": "",
            "stdev": "",
            "min": "",
            "q1": "",
            "q3": "",
            "max": "",
            "pct_gte_80": "",
        }

    sorted_values = sorted(values)
    q1, q3 = _quartiles(sorted_values)
    count_gte_80 = sum(1 for v in sorted_values if v >= 80.0)
    pct_gte_80 = (count_gte_80 / len(sorted_values)) * 100.0 if sorted_values else 0.0
    return {
        "count": len(sorted_values),
        "mean": round(statistics.mean(sorted_values), 4),
        "median": round(statistics.median(sorted_values), 4),
        "stdev": round(statistics.stdev(sorted_values), 4) if len(sorted_values) > 1 else 0.0,
        "min": round(sorted_values[0], 4),
        "q1": round(q1, 4),
        "q3": round(q3, 4),
        "max": round(sorted_values[-1], 4),
        "pct_gte_80": round(pct_gte_80, 2),
    }


def _add_histogram_section(sheet, start_row: int, title: str, values: list[float]) -> int:
    sheet.cell(row=start_row, column=1).value = title
    start_row += 1
    sheet.cell(row=start_row, column=1).value = "bin"
    sheet.cell(row=start_row, column=2).value = "count"
    start_row += 1

    if not values:
        sheet.cell(row=start_row, column=1).value = "(no data)"
        sheet.cell(row=start_row, column=2).value = 0
        return start_row + 1

    sorted_values = sorted(values)
    min_value = sorted_values[0]
    max_value = sorted_values[-1]

    bin_count = 10
    if math.isclose(min_value, max_value):
        bin_edges = [min_value, max_value + 1.0]
    else:
        width = (max_value - min_value) / bin_count
        bin_edges = [min_value + (i * width) for i in range(bin_count)]
        bin_edges.append(max_value)

    frequencies = [0] * (len(bin_edges) - 1)
    for value in sorted_values:
        for idx in range(len(bin_edges) - 1):
            upper_inclusive = idx == len(bin_edges) - 2
            lower = bin_edges[idx]
            upper = bin_edges[idx + 1]
            if (value >= lower) and (value <= upper if upper_inclusive else value < upper):
                frequencies[idx] += 1
                break

    first_data_row = start_row
    for idx, freq in enumerate(frequencies):
        lower = bin_edges[idx]
        upper = bin_edges[idx + 1]
        sheet.cell(row=start_row, column=1).value = f"{lower:.2f} - {upper:.2f}"
        sheet.cell(row=start_row, column=2).value = freq
        start_row += 1

    chart = BarChart()
    chart.type = "col"
    chart.title = "Histogram of submission_numeric_score"
    chart.y_axis.title = "count"
    chart.x_axis.title = "bin"
    data_ref = Reference(sheet, min_col=2, min_row=first_data_row - 1, max_row=start_row - 1)
    cats_ref = Reference(sheet, min_col=1, min_row=first_data_row, max_row=start_row - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 7
    chart.width = 12
    sheet.add_chart(chart, f"D{first_data_row}")

    return start_row + 1


def _add_normalized_histogram_section(
    sheet,
    start_row: int,
    title: str,
    series_values: list[tuple[str, list[float]]],
) -> tuple[int, int, int]:
    sheet.cell(row=start_row, column=1).value = title
    start_row += 1
    sheet.cell(row=start_row, column=1).value = "bin"
    for index, (series_name, _values) in enumerate(series_values, start=2):
        sheet.cell(row=start_row, column=index).value = series_name
    start_row += 1

    bucket_bounds = [(0, 50)] + [(lower, lower + 5) for lower in range(50, 100, 5)]
    bucket_labels = [f"{lower}-{upper}" for lower, upper in bucket_bounds]
    frequencies_by_series: list[list[int]] = []
    for _series_name, values in series_values:
        frequencies = [0] * len(bucket_labels)
        if values:
            for value in values:
                # Clamp to [0, 100] so all observations fit the requested normalized buckets.
                normalized = min(max(value, 0.0), 100.0)
                if normalized <= 50.0:
                    frequencies[0] += 1
                    continue
                bucket_index = int((normalized - 50.0 - 1e-9) // 5.0) + 1
                bucket_index = max(1, min(bucket_index, len(frequencies) - 1))
                frequencies[bucket_index] += 1
        frequencies_by_series.append(frequencies)

    first_data_row = start_row
    for row_offset, label in enumerate(bucket_labels):
        label_cell = sheet.cell(row=start_row, column=1)
        # Keep bucket labels as text so Excel category axis uses explicit string bins.
        label_cell.value = str(label)
        label_cell.number_format = "@"
        for series_index, frequencies in enumerate(frequencies_by_series, start=2):
            sheet.cell(row=start_row, column=series_index).value = frequencies[row_offset]
        start_row += 1

    last_data_row = start_row - 1
    return start_row + 1, first_data_row, last_data_row


def write_xlsx(
    output_path: Path,
    ordered_identity_keys: list[tuple[str, str, str]],
    source_data: list[tuple[str, dict[tuple[str, str, str], dict[str, str]], str, str]],
    component_weights: dict[str, str],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "component_grades"
    normalized_summary_scores: list[float] = []
    normalized_weighted_component_scores: dict[str, list[float]] = {
        source_name: [] for source_name, _rows_by_id, _feedback_column, _max_grade_column in source_data
    }

    header = list(IDENTITY_COLUMNS)
    weighted_grade_headers: list[str] = []
    weighted_max_headers: list[str] = []
    weighted_weight_headers: list[str] = []
    scaled_feedback_headers: list[str] = []
    for index, (source_name, _rows_by_id, _feedback_column, _max_grade_column) in enumerate(source_data):
        if index > 0:
            header.append(SEPARATOR_COLUMN)
        header.append(f"Raw_Grade_{source_name}")
        header.append(f"Raw_Max_{source_name}")
        header.append(f"Feedback_{source_name}")
        weighted_weight_headers.append(f"Weight_{source_name}")
        scaled_feedback_headers.append(f"Feedback_Scaled_{source_name}")
        weighted_grade_headers.append(f"Weighted_Grade_{source_name}")
        weighted_max_headers.append(f"Weighted_Max_{source_name}")
    header.append(SEPARATOR_COLUMN)
    for index, (weight_header, scaled_feedback_header, grade_header, max_header) in enumerate(zip(
        weighted_weight_headers,
        scaled_feedback_headers,
        weighted_grade_headers,
        weighted_max_headers,
    )):
        if index > 0:
            header.append(SEPARATOR_COLUMN)
        header.append(weight_header)
        header.append(scaled_feedback_header)
        header.append(grade_header)
        header.append(max_header)
    header.append(SEPARATOR_COLUMN)
    header.append(WEIGHTED_SCORE_COLUMN)
    header.append(MAX_SCORE_COLUMN)
    header.append(AGGREGATED_FEEDBACK_COLUMN)
    sheet.append(header)

    weighted_score_column_index = header.index(WEIGHTED_SCORE_COLUMN)
    weighted_max_column_index = header.index(MAX_SCORE_COLUMN)
    raw_grade_column_indexes = [
        index
        for index, column_name in enumerate(header)
        if isinstance(column_name, str) and column_name.startswith("Raw_Grade_")
    ]
    raw_max_column_indexes = [
        index
        for index, column_name in enumerate(header)
        if isinstance(column_name, str) and column_name.startswith("Raw_Max_")
    ]
    weight_column_indexes = [
        index
        for index, column_name in enumerate(header)
        if isinstance(column_name, str) and column_name.startswith("Weight_")
    ]
    weighted_grade_column_indexes = [
        index
        for index, column_name in enumerate(header)
        if isinstance(column_name, str) and column_name.startswith("Weighted_Grade_")
    ]
    weighted_max_column_indexes = [
        index
        for index, column_name in enumerate(header)
        if isinstance(column_name, str) and column_name.startswith("Weighted_Max_")
    ]
    scaled_feedback_column_indexes = [
        index
        for index, column_name in enumerate(header)
        if isinstance(column_name, str) and column_name.startswith("Feedback_Scaled_")
    ]
    aggregated_feedback_column_index = header.index(AGGREGATED_FEEDBACK_COLUMN)

    for key in ordered_identity_keys:
        row_out = list(key)
        raw_grade_values: list[str] = []
        raw_max_values: list[str] = []
        weight_values: list[str] = []
        scaled_feedback_values: list[str] = []
        for index, (_source_name, rows_by_id, feedback_column, max_grade_column) in enumerate(source_data):
            if index > 0:
                row_out.append("")
            source_row = rows_by_id[key]
            grade_value_raw = (source_row.get(GRADE_COLUMN, "") or "").strip()
            max_grade_value_raw = (source_row.get(max_grade_column, "") or "").strip()
            weight_value = component_weights.get(_source_name, "")
            feedback_text_raw = (source_row.get(feedback_column, "") or "").strip()
            feedback_text_scaled = scale_feedback_score_pairs(feedback_text_raw, weight_value)
            parsed_grade = _parse_float(grade_value_raw)
            parsed_max_grade = _parse_float(max_grade_value_raw)
            row_out.append(parsed_grade if parsed_grade is not None else grade_value_raw)
            row_out.append(parsed_max_grade if parsed_max_grade is not None else max_grade_value_raw)
            row_out.append(feedback_text_raw)
            raw_grade_values.append(grade_value_raw)
            raw_max_values.append(max_grade_value_raw)
            weight_values.append(weight_value)
            scaled_feedback_values.append(feedback_text_scaled)
        row_out.append("")
        for index, (weight_value, grade_value, max_grade_value, scaled_feedback_value) in enumerate(zip(
            weight_values,
            raw_grade_values,
            raw_max_values,
            scaled_feedback_values,
        )):
            if index > 0:
                row_out.append("")
            parsed_weight = _parse_float(weight_value)
            row_out.append(parsed_weight if parsed_weight is not None else weight_value)
            row_out.append(scaled_feedback_value)
            row_out.append("")
            row_out.append("")
        row_out.append("")
        row_out.append("")
        row_out.append("")
        row_out.append("")
        sheet.append(row_out)

        weighted_total_score = _compute_weighted_total(raw_grade_values, weight_values)
        weighted_total_max = _compute_weighted_total(raw_max_values, weight_values)
        if (
            weighted_total_score is not None
            and weighted_total_max is not None
            and not math.isclose(weighted_total_max, 0.0)
        ):
            normalized_summary_scores.append((weighted_total_score / weighted_total_max) * 100.0)

        for source_name, raw_grade_value, raw_max_value, weight_value in zip(
            [name for name, _rows_by_id, _feedback_column, _max_grade_column in source_data],
            raw_grade_values,
            raw_max_values,
            weight_values,
        ):
            grade = _parse_float(raw_grade_value)
            max_grade = _parse_float(raw_max_value)
            weight = _parse_float(weight_value)
            if grade is None or max_grade is None or weight is None:
                continue
            weighted_grade = grade * weight
            weighted_max = max_grade * weight
            if math.isclose(weighted_max, 0.0):
                continue
            normalized_weighted_component_scores[source_name].append((weighted_grade / weighted_max) * 100.0)

        sheet_row = sheet.max_row
        for idx, (weight_idx, raw_grade_idx, weighted_grade_idx) in enumerate(
            zip(weight_column_indexes, raw_grade_column_indexes, weighted_grade_column_indexes)
        ):
            weight_ref = f"{get_column_letter(weight_idx + 1)}{sheet_row}"
            raw_grade_ref = f"{get_column_letter(raw_grade_idx + 1)}{sheet_row}"
            weighted_grade_ref_col = weighted_grade_idx + 1
            sheet.cell(row=sheet_row, column=weighted_grade_ref_col).value = (
                f'=IF(OR({weight_ref}="",{raw_grade_ref}=""),"",{weight_ref}*{raw_grade_ref})'
            )

        for idx, (weight_idx, raw_max_idx, weighted_max_idx) in enumerate(
            zip(weight_column_indexes, raw_max_column_indexes, weighted_max_column_indexes)
        ):
            weight_ref = f"{get_column_letter(weight_idx + 1)}{sheet_row}"
            raw_max_ref = f"{get_column_letter(raw_max_idx + 1)}{sheet_row}"
            weighted_max_ref_col = weighted_max_idx + 1
            sheet.cell(row=sheet_row, column=weighted_max_ref_col).value = (
                f'=IF(OR({weight_ref}="",{raw_max_ref}=""),"",{weight_ref}*{raw_max_ref})'
            )

        weighted_grade_refs = ",".join(
            f"{get_column_letter(column_index + 1)}{sheet_row}"
            for column_index in weighted_grade_column_indexes
        )
        sheet.cell(row=sheet_row, column=weighted_score_column_index + 1).value = (
            f"=IF(COUNTA({weighted_grade_refs})=0,\"\",SUM({weighted_grade_refs}))"
        )

        weighted_max_refs = ",".join(
            f"{get_column_letter(column_index + 1)}{sheet_row}"
            for column_index in weighted_max_column_indexes
        )
        sheet.cell(row=sheet_row, column=weighted_max_column_index + 1).value = (
            f"=IF(COUNTA({weighted_max_refs})=0,\"\",SUM({weighted_max_refs}))"
        )
        if scaled_feedback_column_indexes:
            feedback_refs = [
                f"{get_column_letter(column_index + 1)}{sheet_row}"
                for column_index in scaled_feedback_column_indexes
            ]
            # Build a compatibility-safe concatenation formula that avoids TEXTJOIN,
            # which some Excel builds rewrite as @TEXTJOIN and then fail with #NAME?.
            formula_terms: list[str] = []
            for idx, ref in enumerate(feedback_refs):
                if idx == 0:
                    formula_terms.append(f'IF({ref}<>"",{ref},"")')
                    continue
                previous_refs = feedback_refs[:idx]
                previous_non_empty = "OR(" + ",".join(f'{prev}<>""' for prev in previous_refs) + ")"
                formula_terms.append(
                    f'IF({ref}<>"",IF({previous_non_empty},CHAR(10)&CHAR(10),"")&{ref},"")'
                )
            formula_expr = "&".join(formula_terms)
            sheet.cell(row=sheet_row, column=aggregated_feedback_column_index + 1).value = (
                f"={formula_expr}"
            )

    summary_sheet = workbook.create_sheet(title="summary_stats")
    summary_series_values: list[tuple[str, list[float]]] = [
        ("submission_numeric_score_norm100", normalized_summary_scores)
    ]
    for source_name, _rows_by_id, _feedback_column, _max_grade_column in source_data:
        summary_series_values.append(
            (f"Weighted_Grade_{source_name}_norm100", normalized_weighted_component_scores[source_name])
        )

    summary_sheet.cell(row=1, column=1).value = "metric"
    for col_index, (series_name, _values) in enumerate(summary_series_values, start=2):
        summary_sheet.cell(row=1, column=col_index).value = series_name

    metrics_order = ["count", "mean", "median", "stdev", "min", "q1", "q3", "max", "pct_gte_80"]
    for row_index, metric_name in enumerate(metrics_order, start=2):
        # Map metric names to friendly display labels
        display_name = "% A/A+ (≥80)" if metric_name == "pct_gte_80" else metric_name
        summary_sheet.cell(row=row_index, column=1).value = display_name
        for col_index, (_series_name, values) in enumerate(summary_series_values, start=2):
            metric_values = _metric_row_values(values)
            summary_sheet.cell(row=row_index, column=col_index).value = metric_values[metric_name]

    histogram_data_sheet = workbook.create_sheet(title="histogram_data")
    histogram_series_values = summary_series_values
    _next_row, first_data_row, last_data_row = _add_normalized_histogram_section(
        histogram_data_sheet,
        1,
        "Histogram bins for normalized submission and weighted component grades (0-100)",
        histogram_series_values,
    )
    histogram_data_sheet.sheet_state = "visible"

    histogram_chart = BarChart()
    histogram_chart.type = "col"
    histogram_chart.title = "Histogram of normalized submission grade (0-100)"
    histogram_chart.y_axis.title = "count"
    histogram_chart.x_axis.title = "Normalized grade bucket (0-100)"
    histogram_chart.x_axis.delete = False
    histogram_chart.y_axis.delete = False
    histogram_chart.x_axis.tickLblPos = "low"
    histogram_chart.legend = None
    # Ensure category labels still render when source data lives on a hidden sheet.
    histogram_chart.visible_cells_only = False
    data_ref = Reference(
        histogram_data_sheet,
        min_col=2,
        min_row=first_data_row - 1,
        max_row=last_data_row,
    )
    cats_ref = Reference(
        histogram_data_sheet,
        min_col=1,
        min_row=first_data_row,
        max_row=last_data_row,
    )
    histogram_chart.add_data(data_ref, titles_from_data=True)
    histogram_chart.set_categories(cats_ref)
    category_formula = f"'{histogram_data_sheet.title}'!$A${first_data_row}:$A${last_data_row}"
    for series in histogram_chart.series:
        series.cat = AxDataSource(strRef=StrRef(f=category_formula))
        # Use a single color for all histogram bars
        series.graphicalProperties.solidFill = "4472C4"  # Standard blue

    histogram_chart_sheet = workbook.create_chartsheet(title="histogram")
    histogram_chart_sheet.add_chart(histogram_chart)

    # Create a second histogram chart showing all components with different colors
    component_color_palette = [
        "4472C4",  # Blue
        "ED7D31",  # Orange
        "A5A5A5",  # Gray
        "FFC000",  # Gold
        "5B9BD5",  # Light Blue
        "70AD47",  # Green
        "FF6B6B",  # Red
        "6F42C1",  # Purple
        "20C997",  # Teal
        "FFC107",  # Amber
        "17A2B8",  # Cyan
        "E83E8C",  # Pink
        "6C757D",  # Dark Gray
        "28A745",  # Dark Green
        "DC3545",  # Dark Red
    ]

    component_histogram_chart = BarChart()
    component_histogram_chart.type = "col"
    component_histogram_chart.title = "Histogram by Assignment Component"
    component_histogram_chart.y_axis.title = "count"
    component_histogram_chart.x_axis.title = "Normalized grade bucket (0-100)"
    component_histogram_chart.x_axis.delete = False
    component_histogram_chart.y_axis.delete = False
    component_histogram_chart.x_axis.tickLblPos = "low"
    component_histogram_chart.legend.position = "r"  # Right legend
    component_histogram_chart.visible_cells_only = False

    # Add each component series with its own color (skip the first submission score series)
    for series_index, (series_name, _values) in enumerate(histogram_series_values[1:], start=2):
        series_data_ref = Reference(
            histogram_data_sheet,
            min_col=series_index,
            min_row=first_data_row - 1,
            max_row=last_data_row,
        )
        component_histogram_chart.add_data(series_data_ref, titles_from_data=True)
        
        # Get the most recently added series and set its color
        color_index = (series_index - 2) % len(component_color_palette)
        series_obj = component_histogram_chart.series[-1]
        series_obj.graphicalProperties.solidFill = component_color_palette[color_index]

    # Set categories for the component histogram
    cats_ref = Reference(
        histogram_data_sheet,
        min_col=1,
        min_row=first_data_row,
        max_row=last_data_row,
    )
    component_histogram_chart.set_categories(cats_ref)
    category_formula = f"'{histogram_data_sheet.title}'!$A${first_data_row}:$A${last_data_row}"
    for series in component_histogram_chart.series:
        series.cat = AxDataSource(strRef=StrRef(f=category_formula))

    component_histogram_chart_sheet = workbook.create_chartsheet(title="histogram_by_component")
    component_histogram_chart_sheet.add_chart(component_histogram_chart)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_template_csv(
    output_path: Path,
    iteration: str,
    ordered_identity_keys: list[tuple[str, str, str]],
    source_data: list[tuple[str, dict[tuple[str, str, str], dict[str, str]], str, str]],
    component_weights: dict[str, str],
    is_ssid: bool = False,
    token: str = "",
) -> None:
    """Write a gradebook CSV with stage02-like populated summary columns.
    
    If is_ssid=True, relabels Identifier to SSID and adds a normalized grade column
    named after the token (assignment name).
    """
    identifier_column = "SSID" if is_ssid else "Identifier"
    normalized_grade_column = token if is_ssid else "grade_normalized"
    header = [
        identifier_column,
        "Full name",
        "Email address",
        "Status",
        "Grade",
        "Maximum Grade",
    ]
    if is_ssid:
        header.append(normalized_grade_column)
    header.extend([
        "Grade can be changed",
        "Last modified (submission)",
        "Online text",
        "Last modified (grade)",
        "Feedback comments",
    ])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=header)
        writer.writeheader()
        for key in ordered_identity_keys:
            total_grade = 0.0
            total_max_grade = 0.0
            grade_terms = 0
            max_grade_terms = 0
            for source_name, rows_by_id, _feedback_column, max_grade_column in source_data:
                source_row = rows_by_id[key]
                weight_value = component_weights.get(source_name, "")

                scaled_grade = scale_value_by_weight((source_row.get(GRADE_COLUMN, "") or "").strip(), weight_value)
                if scaled_grade is not None:
                    total_grade += scaled_grade
                    grade_terms += 1

                scaled_max_grade = scale_value_by_weight(
                    (source_row.get(max_grade_column, "") or "").strip(),
                    weight_value,
                )
                if scaled_max_grade is not None:
                    total_max_grade += scaled_max_grade
                    max_grade_terms += 1

            grade_str = f"{total_grade:.2f}" if grade_terms > 0 else ""
            max_grade_str = f"{total_max_grade:.2f}" if max_grade_terms > 0 else ""
            
            # Calculate normalized grade
            grade_normalized = ""
            if is_ssid and grade_str and max_grade_str:
                try:
                    grade_val = float(grade_str)
                    max_val = float(max_grade_str)
                    if max_val > 0:
                        grade_normalized = f"{(grade_val / max_val):.4f}"
                except (ValueError, ZeroDivisionError):
                    grade_normalized = ""

            row_out = {
                identifier_column: key[0],
                "Full name": key[1],
                "Email address": key[2],
                "Status": "",
                "Grade": grade_str,
                "Maximum Grade": max_grade_str,
                "Grade can be changed": "",
                "Last modified (submission)": "",
                "Online text": "",
                "Last modified (grade)": "",
                "Feedback comments": aggregate_feedback_for_identity(key, source_data, component_weights),
            }
            if is_ssid:
                row_out[normalized_grade_column] = grade_normalized
            writer.writerow(row_out)


def build_source_data(
    discovered: list[tuple[str, Path]],
) -> tuple[
    list[tuple[str, dict[tuple[str, str, str], dict[str, str]], str, str]],
    list[tuple[str, str, str]],
]:
    source_data: list[tuple[str, dict[tuple[str, str, str], dict[str, str]], str, str]] = []
    baseline_keys: set[tuple[str, str, str]] | None = None
    ordered_identity_keys: list[tuple[str, str, str]] = []

    for source_name, csv_path in discovered:
        _headers, rows, feedback_column, max_grade_column = read_grade_rows(csv_path)
        rows_by_id = validate_and_index_identity_rows(source_name, rows)
        current_keys = set(rows_by_id.keys())

        if baseline_keys is None:
            baseline_keys = current_keys
            ordered_identity_keys = list(rows_by_id.keys())
        else:
            if current_keys != baseline_keys:
                missing_from_current = sorted(baseline_keys - current_keys)
                extra_in_current = sorted(current_keys - baseline_keys)
                raise ValueError(
                    "Identity precondition failed for source "
                    f"{source_name}. Missing keys: {missing_from_current[:3]} "
                    f"Extra keys: {extra_in_current[:3]}"
                )
            # Verify identity text values are identical across sources.
            for key in baseline_keys:
                baseline_identifier, baseline_name, baseline_email = key
                current_identifier, current_name, current_email = key
                if (
                    baseline_identifier != current_identifier
                    or baseline_name != current_name
                    or baseline_email != current_email
                ):
                    raise ValueError(f"Identity values mismatch for key {key} in source {source_name}")

        source_data.append((source_name, rows_by_id, feedback_column, max_grade_column))

    return source_data, ordered_identity_keys


def main() -> int:
    args = parse_args()
    assignment_root = args.assignment_root.resolve()
    if not assignment_root.is_dir():
        raise FileNotFoundError(f"Assignment root not found: {assignment_root}")

    token = assignment_root.name
    grades_release_dir = assignment_root / "04_grades_release"
    output_xlsx = grades_release_dir / f"{token}_component_grade_feedback_merged.xlsx"
    weights_path = grades_release_dir / f"{token}_weights.json"
    component_weights = load_component_weights(weights_path)
    weighted_components = set(component_weights.keys())

    assignment_dirs = sorted(
        [path for path in assignment_root.iterdir() if path.is_dir()],
        key=lambda path: path.name,
    )

    discovered: list[tuple[str, Path]] = []
    discovered_ssid: list[tuple[str, Path]] = []
    skipped_missing_csv: list[tuple[str, Path]] = []
    skipped_missing_ssid_csv: list[tuple[str, Path]] = []
    first_config: dict[str, object] | None = None
    first_iteration = ""
    
    assignment_dir_by_name = {path.name: path for path in assignment_dirs}
    missing_component_dirs = sorted(weighted_components - set(assignment_dir_by_name.keys()))
    if missing_component_dirs:
        raise ValueError(
            "Weights file lists components that are missing under assignment root: "
            f"{missing_component_dirs}"
        )

    for component_name in sorted(weighted_components):
        assignment_dir = assignment_dir_by_name[component_name]
        config = load_pipeline_paths_config(assignment_dir)
        if config is None:
            raise ValueError(
                f"Missing pipeline_paths.json for weighted component: {assignment_dir}"
            )
        if first_config is None:
            first_config = config
            layer4_prod = config.get("layer4_prod", {})
            if isinstance(layer4_prod, dict):
                release = layer4_prod.get("release", {})
                if isinstance(release, dict):
                    first_iteration = str(release.get("iteration", "")).strip()
        csv_path = resolve_output_csv_from_config(assignment_dir, config)
        if not csv_path.exists() or not csv_path.is_file():
            raise FileNotFoundError(
                f"Missing required gradebook CSV for weighted component {assignment_dir.name}: {csv_path}"
            )
        discovered.append((assignment_dir.name, csv_path))

        ssid_csv_path = csv_path.with_name(f"{csv_path.stem}_SSID{csv_path.suffix}")
        if ssid_csv_path.exists() and ssid_csv_path.is_file():
            discovered_ssid.append((assignment_dir.name, ssid_csv_path))
        else:
            skipped_missing_ssid_csv.append((assignment_dir.name, ssid_csv_path))

    if not discovered:
        raise ValueError(f"No assignment subdirectories with pipeline_paths.json found under {assignment_root}")

    source_data, ordered_identity_keys = build_source_data(discovered)

    write_xlsx(output_xlsx, ordered_identity_keys, source_data, component_weights)
    
    # Write template CSV with standard gradebook submission columns
    template_filename = f"{token}_Grades_iter{first_iteration}.csv"
    template_path = grades_release_dir / template_filename
    write_template_csv(template_path, first_iteration, ordered_identity_keys, source_data, component_weights, token=token)

    ssid_output_xlsx: Path | None = None
    ssid_template_path: Path | None = None
    ssid_source_data: list[tuple[str, dict[tuple[str, str, str], dict[str, str]], str, str]] = []
    ssid_ordered_identity_keys: list[tuple[str, str, str]] = []
    if discovered_ssid:
        ssid_source_data, ssid_ordered_identity_keys = build_source_data(discovered_ssid)
        ssid_output_xlsx = grades_release_dir / f"{token}_component_grade_feedback_merged_SSID.xlsx"
        write_xlsx(ssid_output_xlsx, ssid_ordered_identity_keys, ssid_source_data, component_weights)

        ssid_template_filename = f"{token}_Grades_iter{first_iteration}_SSID.csv"
        ssid_template_path = grades_release_dir / ssid_template_filename
        write_template_csv(
            ssid_template_path,
            first_iteration,
            ssid_ordered_identity_keys,
            ssid_source_data,
            component_weights,
            is_ssid=True,
            token=token,
        )

    print(f"ASST_ROOT: {assignment_root}")
    print(f"Grades release dir: {grades_release_dir}")
    print(f"Token: {token}")
    print(f"Iteration: {first_iteration}")
    print(f"Input sources found: {len(source_data)}")
    for source_name, csv_path in discovered:
        print(f"- {source_name}: {csv_path}")
    if skipped_missing_csv:
        print(f"Skipped missing CSVs: {len(skipped_missing_csv)}")
        for source_name, csv_path in skipped_missing_csv:
            print(f"- {source_name}: {csv_path}")
    print(f"SSID input sources found: {len(discovered_ssid)}")
    for source_name, csv_path in discovered_ssid:
        print(f"- {source_name}: {csv_path}")
    if skipped_missing_ssid_csv:
        print(f"Skipped missing SSID CSVs: {len(skipped_missing_ssid_csv)}")
        for source_name, csv_path in skipped_missing_ssid_csv:
            print(f"- {source_name}: {csv_path}")
    print(f"Rows merged: {len(ordered_identity_keys)}")
    print(f"Output XLSX: {output_xlsx}")
    print(f"Template CSV: {template_path}")
    if ssid_output_xlsx and ssid_template_path:
        print(f"SSID Rows merged: {len(ssid_ordered_identity_keys)}")
        print(f"SSID Output XLSX: {ssid_output_xlsx}")
        print(f"SSID Template CSV: {ssid_template_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
